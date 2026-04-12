#!/usr/bin/env python3
"""
QC Dashboard — 数据采集器
===========================
支持三种模式从企业微信智能表格提取数据：
  1. wecom-api: 通过 wecom-cli API 自动拉取（推荐，需配置凭证）
  2. cdp:         通过 Chrome DevTools Protocol 浏览器自动化提取
  3. excel:       解析手动导出的 Excel 文件

用法:
  python src/collector.py setup          # 初始化 wecom-cli 凭证（首次必须）
  python src/collector.py fetch          # 拉取所有队列数据
  python src/collector.py fetch --queue q6_shangqiang   # 只拉取指定队列
  python src/collector.py import         # 导入 Excel 数据
  python src/collector.py status         # 查看各队列数据状态
  python src/collector.py export-json    # 导出 JSON 数据
  python src/collector.py export-html    # 导出带数据的完整 HTML 看板
"""

import os, sys, json, yaml, logging, subprocess, shutil
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional

# ── 统一数据库连接（自动处理 SQLCipher 加密）──
from db_helper import get_db_connection, _get_password as _get_db_password

# ── 项目根目录 ──
PROJECT_ROOT = Path(__file__).parent.parent
CONFIG_FILE = PROJECT_ROOT / "config.yaml"
DB_FILE = PROJECT_ROOT / "data" / "metrics.db"
DATA_DIR = PROJECT_ROOT / "data"
EXCEL_INPUT_DIR = PROJECT_ROOT / "data" / "uploads"
DASHBOARD_HTML = PROJECT_ROOT / "dashboard" / "index.html"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger("qc-collector")


# ============================================================
# 配置加载
# ============================================================
def load_config() -> dict:
    """加载 config.yaml"""
    if not CONFIG_FILE.exists():
        log.error(f"配置文件不存在: {CONFIG_FILE}")
        log.info("请复制 config.yaml.example 为 config.yaml 并填写配置")
        sys.exit(1)
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def get_wecom_cli_path() -> Optional[str]:
    """查找 wecom-cli 可执行文件路径"""
    path = shutil.which('wecom-cli')
    if path and os.path.isfile(path):
        return path
    fallback = os.path.expanduser('~/.local/bin/wecom-cli')
    if os.path.isfile(fallback):
        return fallback
    return None


def check_wecom_init() -> bool:
    """检查 wecom-cli 是否已完成 init"""
    cli = get_wecom_cli_path()
    if not cli:
        return False
    # 尝试调用一个简单命令看是否报 "未找到 MCP 配置缓存" 错误
    r = subprocess.run([cli, 'doc', '--help'], capture_output=True, text=True, timeout=10)
    if '未找到 MCP 配置缓存' in r.stderr or '未找到 MCP 配置缓存' in r.stdout:
        return False
    return True


# ============================================================
# SQLite 存储
# ============================================================
def init_db():
    """初始化数据库表"""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS daily_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            queue_id TEXT NOT NULL,
            queue_name TEXT DEFAULT '',
            date TEXT NOT NULL,
            metric_key TEXT NOT NULL,
            metric_value REAL,
            raw_data TEXT,
            source TEXT DEFAULT 'manual',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(queue_id, date, metric_key)
        )
    ''')
    c.execute('CREATE INDEX IF NOT EXISTS idx_qd ON daily_metrics(queue_id, date)')
    conn.commit()
    return conn


def upsert_metrics(conn, queue_id: str, queue_name: str, date_str: str, metrics: dict, source: str = 'manual'):
    """插入或更新一条日期的指标数据"""
    import pandas as pd
    c = conn.cursor()
    for key, value in metrics.items():
        if value is None:
            continue
        
        # 处理 pandas Timestamp / NaT 等特殊类型
        if pd is not None:
            if isinstance(value, pd.Timestamp):
                value = value.strftime('%Y-%m-%d')
            elif pd.isna(value):
                continue
        
        try:
            val = float(value) if not isinstance(value, (int, float)) else float(value)
            # 排除 NaN 和 Inf
            if val != val or val == float('inf') or val == float('-inf'):
                continue
        except (ValueError, TypeError):
            val = str(value)
        
        # 限制 raw_data 大小避免超长
        try:
            raw_json = json.dumps(metrics, ensure_ascii=False)
        except (TypeError, ValueError):
            raw_json = str(metrics)[:2000]
        
        c.execute('''
            INSERT INTO daily_metrics (queue_id, queue_name, date, metric_key, metric_value, raw_data, source, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now','localtime'))
            ON CONFLICT(queue_id, date, metric_key)
            DO UPDATE SET metric_value=excluded.metric_value, raw_data=excluded.raw_data,
                              source=excluded.source, updated_at=datetime('now','localtime')
        ''', (queue_id, queue_name, date_str, key, val, raw_json, source))
    conn.commit()


def get_queue_dates(conn, queue_id: str) -> list:
    """获取某队列已有数据日期列表"""
    c = conn.cursor()
    c.execute("SELECT DISTINCT date FROM daily_metrics WHERE queue_id=? ORDER BY date", (queue_id,))
    return [r[0] for r in c.fetchall()]


def export_as_json(conn) -> dict:
    """导出全部数据为 JSON 结构（供 HTML 使用）"""
    c = conn.cursor()
    c.execute('SELECT DISTINCT queue_id, queue_name FROM daily_metrics')
    queues = {}
    for qid, qname in c.fetchall():
        c2 = conn.cursor()
        c2.execute('SELECT date, metric_key, metric_value FROM daily_metrics WHERE queue_id=? ORDER BY date', (qid,))
        rows = {}
        for date, key, val in c2.fetchall():
            if date not in rows:
                rows[date] = {'_queueName': qname or '', 'date': date}
            rows[date][key] = val
        queues[qid] = sorted(rows.values(), key=lambda x: x['date'])
    return queues


# ============================================================
# 模式1: Excel 导入（保持不变）
# ============================================================
def import_excel(conn, config: dict, queue_id: Optional[str] = None, file_path: Optional[str] = None):
    """从 data/uploads/ 目录或指定文件导入 Excel（使用 pandas + calamine 引擎）
    
    Args:
        conn: 数据库连接
        config: 配置字典
        queue_id: 可选，只导入指定队列
        file_path: 可选，直接指定要导入的文件路径（跳过目录扫描）
    """
    if file_path:
        # 直接使用指定文件
        excel_files = [Path(file_path)]
    else:
        EXCEL_INPUT_DIR.mkdir(parents=True, exist_ok=True)
        excel_files = list(EXCEL_INPUT_DIR.glob("*.xlsx")) + list(EXCEL_INPUT_DIR.glob("*.xls"))
    if not excel_files:
        log.warning(f"未找到 Excel 文件。请将导出的企微表格放到: {EXCEL_INPUT_DIR}")
        log.info("支持格式: .xlsx / .xls")
        return 0

    # 尝试使用 calamine 引擎（兼容性更好）
    try:
        import pandas as pd
        USE_PANDAS = True
    except ImportError:
        log.warning("pandas 未安装。尝试回退到 openpyxl...")
        USE_PANDAS = False
        try:
            import openpyxl
        except ImportError:
            log.error("需要 pandas 或 openpyxl 库")
            return 0

    queues_config = config.get('queues', [])
    imported = 0

    for fpath in excel_files:
        log.info(f"📂 处理文件: {fpath.name}")
        
        # ── 阶段1：匹配队列 ──
        # 优先使用 queue_id 精确匹配（auto_refresh 传入），
        # 其次按文件名模糊匹配（手动导入场景）
        matched_queue = None
        
        if queue_id:
            # queue_id 精确匹配 — 最可靠的方式
            for qcfg in queues_config:
                if qcfg.get('id') == queue_id:
                    matched_queue = qcfg
                    break
            if matched_queue:
                log.info(f"  队列匹配: {matched_queue.get('name', queue_id)} (精确ID匹配)")
        
        if not matched_queue:
            # 文件名模糊匹配（手动上传 xlsx 时的兜底方案）
            for qcfg in queues_config:
                if qcfg.get('name', '') in fpath.name or qcfg.get('full_name', '') in fpath.name or \
                   qcfg.get('sheet_name', '') in fpath.name:
                    matched_queue = qcfg
                    break

        if not matched_queue:
            log.warning(f"  ⚠️ 无法匹配队列（queue_id={queue_id}, 文件名={fpath.name}）")
            continue

        qid = matched_queue['id']
        qname = matched_queue.get('name', '')
        fields_cfg = matched_queue.get('fields', {})
        date_col = fields_cfg.get('date_col', 'A')
        
        # ── 阶段2：按匹配队列的 sheet_name 精确读取数据 ──
        df = None
        sheet_name = ''
        
        if USE_PANDAS:
            try:
                xl = pd.ExcelFile(str(fpath), engine='calamine')
                
                target_sheet = matched_queue.get('sheet_name', '')
                found_sheet = None
                if target_sheet:
                    for sn in xl.sheet_names:
                        if sn == target_sheet or sn.startswith(target_sheet[:4]) or target_sheet in sn:
                            found_sheet = sn
                            break
                
                if not found_sheet:
                    found_sheet = xl.sheet_names[0] if xl.sheet_names else ''
                    
                sheet_name = found_sheet
                
                if target_sheet and found_sheet != target_sheet:
                    log.warning(f"  ⚠️ 子表 '{target_sheet}' 未精确匹配，使用 '{sheet_name}'")
                else:
                    log.info(f"  工作表: {sheet_name} (共 {len(xl.sheet_names)} 个子表)")
                
                df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
                log.info(f"  数据维度: {df.shape[0]} 行 × {df.shape[1]} 列")
            except Exception as e:
                log.error(f"  calamine 读取失败: {e}")
                continue
        else:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
                sheet_name = wb.sheetnames[0]
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                df = pd.DataFrame(rows)
                wb.close()
            except Exception as e:
                log.error(f"  openpyxl 读取失败: {e}")
                continue

        row_count = 0
        
        # 支持多子表队列（如 q3 四期二审有 GT业务/切片GT业务 两个视图）
        sub_tables_cfg = matched_queue.get('sub_tables', [{}])
        
        for sub_idx, sub_cfg in enumerate(sub_tables_cfg):
            sub_fields = sub_cfg.get('fields', fields_cfg)
            sub_date_col = sub_fields.get('date_col', date_col)
            
            if len(sub_tables_cfg) > 1:
                sub_name = sub_cfg.get('name', f'子表{sub_idx+1}')
                log.info(f"  处理 [{sub_name}] (日期列={sub_date_col})")
            
            # 诊断计数器
            _skip_empty = 0
            _skip_no_date = 0
            _skip_no_metrics = 0

            # 从第3行开始遍历数据（前两行为标题）
            for row_idx in range(2, len(df)):
                row_values = df.iloc[row_idx].tolist()

                if all(pd.isna(v) or v is None or str(v).strip() == '' for v in row_values):
                    _skip_empty += 1
                    continue

                date_val = get_cell_val_pandas(row_values, sub_date_col)
                if not date_val:
                    _skip_no_date += 1
                    continue
                date_str = parse_date(date_val)
                if not date_str:
                    _skip_no_date += 1
                    continue

                metrics = {}
                for mcfg in sub_fields.get('metrics', []):
                    col = mcfg.get('col', '')
                    key = mcfg.get('key', '')
                    val = get_cell_val_pandas(row_values, col)
                    if val is not None:
                        metrics[key] = val

                if metrics:
                    upsert_metrics(conn, qid, qname, date_str, metrics, source='excel')
                    row_count += 1
                else:
                    _skip_no_metrics += 1

            # 子表导入0条时输出诊断信息
            if row_count == 0 and len(df) > 2:
                log.warning(f"  ⚠️ 导入0条！诊断: 总行={len(df)-2}, 空行={_skip_empty}, 无日期={_skip_no_date}, 无指标={_skip_no_metrics}")
                # 输出前3行样本帮助排查
                for sample_idx in range(2, min(5, len(df))):
                    sample_row = df.iloc[sample_idx].tolist()
                    sample_date = get_cell_val_pandas(sample_row, sub_date_col)
                    sample_metrics = {mcfg['col']: get_cell_val_pandas(sample_row, mcfg['col']) for mcfg in sub_fields.get('metrics', [])}
                    log.warning(f"  📋 样本行{sample_idx}: date_col({sub_date_col})={sample_date}, metrics={sample_metrics}")

        log.info(f"  ✅ 导入 {row_count} 条记录 → 队列 [{qname or qid}]")
        imported += row_count

        # 注意：不在这里删除文件！
        # 共享文件（如 q3/q3b 同一 xlsx）需要被多次导入，
        # 由调用方（auto_refresh.py）在全部队列导入完成后统一清理。

    return imported


def get_cell_val(row, col_ref: str):
    """openpyxl 版本：根据列引用（如 A/B/C）从 row tuple 中取值"""
    if not col_ref or col_ref.strip() == '':
        return None
    col_ref = col_ref.upper().strip()
    col_idx = 0
    for ch in col_ref:
        col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
    col_idx -= 1
    if col_idx < len(row):
        v = row[col_idx]
        if v is None or v == '' or str(v).strip() == '':
            return None
        return v
    return None


def get_cell_val_pandas(row_values, col_ref: str):
    """pandas 版本：根据列引用（如 A/B/C/...Z/AA/AB）从 list 中取值
    
    row_values 是 df.iloc[row_idx].tolist() 的结果（有序列的列表）
    """
    import pandas as pd
    if not col_ref or col_ref.strip() == '':
        return None
    col_ref = col_ref.upper().strip()
    
    # 将 Excel 列号转为 0-based 索引 (A=0, B=1, ..., Z=25, AA=26, ...)
    col_idx = 0
    for ch in col_ref:
        if not ('A' <= ch <= 'Z'):
            return None
        col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
    col_idx -= 1

    # 检查索引是否在范围内
    if col_idx < 0 or col_idx >= len(row_values):
        return None
    
    v = row_values[col_idx]
    
    # 处理 pandas 特殊类型
    if pd.isna(v) or v is None:
        return None
    s = str(v).strip()
    if s == '' or s == 'nan' or s == 'NaN' or s == 'NaT':
        return None
    return v


def parse_date(val) -> Optional[str]:
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    formats = ['%Y/%m/%d', '%Y-%m-%d', '%Y.%m.%d', '%Y年%m月%d日',
               '%m/%d/%Y', '%m-%d-%Y']
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    try:
        from datetime import date as d
        days = int(float(s))
        base = d(1899, 12, 30)
        result = base + timedelta(days=days)
        return result.strftime('%Y-%m-%d')
    except:
        pass
    return None


# ============================================================
# 模式2: wecom-cli API（核心模式）
# ============================================================
def setup_wecom(config: dict):
    """
    初始化 wecom-cli 凭证。
    
    流程：
    1. 从 config.yaml 读取 bot_id 和 secret
    2. 调用 wecom-cli init 完成认证
    3. 验证是否可以正常调 API
    """
    wcom_cfg = config.get('global', {}).get('wcom_api', {})
    bot_id = wcom_cfg.get('bot_id', '').strip()
    secret = wcom_cfg.get('secret', '').strip()

    if not bot_id or not secret:
        print("\n" + "=" * 60)
        print("  ⚠️  请先在 config.yaml 中填写凭证信息：")
        print("=" * 60)
        print(f"""
  文件位置: {CONFIG_FILE}
  
  需要填写以下两项：
  
    wcom_api:
      bot_id: ""     ← 在这里填入机器人 ID
      secret: ""     ← 在这里填入应用密钥
  
  获取方式：
    1. 登录企微管理后台 → https://work.weixin.qq.com/wadmin
    2. 应用管理 → 找到/创建「智能机器人」应用
    3. 复制 AgentId 和 Secret 到上方配置中
    
  参考文档：https://open.work.weixin.qq.com/help2/pc/cat?doc_id=21677
""")
        return False

    cli = get_wecom_cli_path()
    if not cli:
        log.error("❌ wecom-cli 未安装。运行: sh install.sh")
        return False

    # 检查是否已经初始化过
    if check_wecom_init():
        log.info("✅ wecom-cli 已完成初始化，跳过")

        # 可选：验证一下能否调通
        log.info("🔍 验证 API 连通性...")
        r = subprocess.run([cli, 'contact', 'get_userlist', '{}'],
                          capture_output=True, text=True, timeout=30)
        if r.returncode == 0:
            log.info("✅ API 连接正常！可以开始拉取数据了。")
        else:
            log.warning(f"⚠️ API 返回异常：{r.stderr[:200] if r.stderr else r.stdout[:200]}")
        return True

    # 需要执行 init
    print("\n" + "=" * 60)
    print("  🔐 正在初始化 wecom-cli 认证...")
    print("=" * 60)
    print(f"\n  Bot ID : {bot_id}")
    print(f"  Secret : {'*' * len(secret)} (已隐藏)")
    print("\n  接下来 wecom-cli 可能会要求交互式输入，请按提示操作。")
    print("  如果需要扫码授权，请在弹出的窗口中完成操作。\n")

    # wecom-cli init 是交互式的，直接运行让用户操作
    r = subprocess.run([cli, 'init'])

    if r.returncode == 0:
        log.info("✅ 初始化成功！")
        # 立即验证
        r2 = subprocess.run([cli, 'contact', 'get_userlist', '{}'],
                           capture_output=True, text=True, timeout=30)
        if r2.returncode == 0:
            log.info("✅ API 验证通过！现在可以运行 `python src/collector.py fetch` 拉取数据了。")
        return True
    else:
        log.error("❌ 初始化失败。请检查 Bot ID 和 Secret 是否正确。")
        return False


def fetch_wecom_api(conn, config: dict, queue_id: Optional[str] = None):
    """通过 wecom-cli API 拉取所有队列数据"""
    # 前置检查
    if not check_wecom_init():
        log.error("❌ wecom-cli 未初始化！请先运行:")
        log.error("   python src/collector.py setup")
        return 0

    cli = get_wecom_cli_path()
    if not cli:
        log.error("❌ wecom-cli 未找到")
        return 0

    queues = config.get('queues', [])
    if queue_id:
        queues = [q for q in queues if q.get('id') == queue_id]
        if not queues:
            log.error(f"找不到队列: {queue_id}")
            return 0

    total = 0
    for qcfg in queues:
        qid = qcfg['id']
        qname = qcfg.get('name', '')
        doc_id = qcfg.get('doc_id', '')
        doc_url = qcfg.get('doc_url', '')

        if not doc_id and not doc_url:
            log.warning(f"  ⏭️ 队列 [{qname}] 缺少文档信息，跳过")
            continue

        log.info(f"📡 拉取队列: [{qname}] ({qid})")

        # ── Step 1: 获取子表列表 ──
        param = json.dumps({'docid': doc_id}) if doc_id else json.dumps({'url': doc_url})
        result = run_cli(cli, 'smartsheet_get_sheet', param)

        if result.get('errcode', -1) != 0:
            log.error(f"  ❌ 获取子表失败: {result.get('errmsg', '未知错误')}")
            log.debug(f"  原始响应: {json.dumps(result, ensure_ascii=False)[:300]}")
            continue

        sheets = result.get('sheets', result.get('sheet_list', []))
        if not sheets:
            log.warning(f"  ⚠️ 该文档没有子表")
            continue

        target_sheet = None
        target_tab = qcfg.get('sheet_tab', '')
        sheet_name_target = qcfg.get('sheet_name', '')

        for sh in sheets:
            sh_title = sh.get('title', '') or sh.get('sheet_title', '')
            sh_id = sh.get('sheet_id', '')
            if sh_title == sheet_name_target or sh_id == target_tab:
                target_sheet = sh
                break

        # 如果没匹配到，取第一个子表
        if not target_sheet and sheets:
            target_sheet = sheets[0]
            log.info(f"  ℹ️ 未精确匹配子表 '{sheet_name_target}'，使用默认: {target_sheet.get('title','')}")

        if not target_sheet:
            log.warning(f"  ⚠️ 所有子表都不匹配，可用: {[s.get('title','') for s in sheets]}")
            continue

        sheet_id = target_sheet.get('sheet_id', '')
        log.info(f"  📋 子表: {target_sheet.get('title','')} (id={sheet_id})")

        # ── Step 2: 获取字段映射（可选，用于列名匹配）───
        field_map = {}  # field_id -> field_title
        fields_result = run_cli(cli, 'smartsheet_get_fields',
                                json.dumps({'docid': doc_id, 'sheet_id': sheet_id}))
        if fields_result.get('errcode', -1) == 0:
            for fld in fields_result.get('fields', []):
                fid = fld.get('field_id', '')
                ftit = fld.get('field_title', '')
                if fid and ftit:
                    field_map[fid] = ftit
                    field_map[ftit] = fid
            log.debug(f"  字段映射: {list(field_map.keys())[:10]}...")

        # ── Step 3: 拉取记录 ──
        record_param = json.dumps({'docid': doc_id, 'sheet_id': sheet_id})
        result = run_cli(cli, 'smartsheet_get_records', record_param)

        if result.get('errcode', -1) != 0:
            log.error(f"  ❌ 拉取记录失败: {result.get('errmsg', '未知错误')}")
            continue

        records = result.get('records', [])

        # 支持多子表（如 q3 四期-二审有 GT业务/切片GT业务 两个视图）
        sub_tables = qcfg.get('sub_tables', [{}])
        
        all_count = 0
        for sub_idx, sub_cfg in enumerate(sub_tables):
            sub_fields = sub_cfg.get('fields', qcfg.get('fields', {}))
            date_col = sub_fields.get('date_col', 'A')

            count = 0
            for rec in records:
                values = rec.get('values', {})

                # 提取日期 — 先尝试列引用，再尝试字段名
                date_val = extract_field_val(values, date_col, field_map)
                if not date_val:
                    continue
                date_str = parse_date(date_val)
                if not date_str:
                    continue

                # 提取指标值
                metrics = {}
                for mcfg in sub_fields.get('metrics', []):
                    key = mcfg.get('key', '')
                    col = mcfg.get('col', '')
                    val = extract_field_val(values, col, field_map)
                    if val is not None:
                        metrics[key] = val

                if metrics:
                    upsert_metrics(conn, qid, qname, date_str, metrics, source='wcom-api')
                    count += 1

            if sub_tables and len(sub_tables) > 1:
                sub_name = sub_cfg.get('name', f'子表{sub_idx+1}')
                log.info(f"  📊 [{sub_name}] {count} 条记录")
            all_count += count

        log.info(f"  ✅ 共 {all_count} 条记录写入 [{qname}]")
        total += all_count

    return total


def extract_field_val(field_values: dict, col_ref: str, field_map: dict = None):
    """
    从 wecom-cli 返回的记录 values 中提取单元格值。
    
    匹配优先级：
    1. 直接用列引用(A/B/C...)作为 key
    2. 用字段标题(打标日期/一检抽检数...)作为 key
    3. 用 field_id 作为 key
    """
    if not field_values or not col_ref:
        return None

    col_ref = col_ref.upper().strip()

    # 优先级1: 直接用列引用
    val = field_values.get(col_ref)
    if val is not None:
        return parse_wecom_cell(val)

    # 优先级2: 用常见中文字段名映射
    col_name_map = {
        'A': ['打标日期', '日期', 'Date', 'date'],
        'B': ['一检抽检数', '抽检数'],
        'C': ['二检不一致'],
        'D': ['二检一致'],
        'E': ['二检结果正常', '违规准确率'],
        'F': ['二检违规', '漏率'],
        'G': ['漏过'],
        'H': ['漏过删除后', '漏过(删)'],
        'I': ['误伤'],
        'J': ['误伤删除后', '误伤(删)'],
        'K': ['错罚'],
        'L': ['错罚删除后', '错罚(删)'],
        'M': ['一致错误数', '一致错误'],
        'N': ['一致错误数(删除)', '一致错误(删)'],
        'Z': ['审核准确率', '准确率'],
        # 四期-举报专用
        'J': ['申诉前-违规准确率'],
        'M': ['申诉前-漏率'],
        'F': ['申诉前-准确率'],
        'R': ['申诉后-违规准确率'],
        'U': ['申诉后-漏率'],
        'AB': ['申诉后-准确率'],
    }

    names_to_try = col_name_map.get(col_ref, [col_ref])
    for name in names_to_try:
        val = field_values.get(name)
        if val is not None:
            return parse_wecom_cell(val)

    # 优先级3: 用 field_map 反查
    if field_map:
        for fk, fv in field_map.items():
            if col_ref in (fk, fv) or (fv and col_ref in str(fv)):
                val = field_values.get(fk)
                if val is not None:
                    return parse_wecom_cell(val)
                val = field_values.get(fv)
                if val is not None:
                    return parse_wecom_cell(val)

    return None


def parse_wecom_cell(cell_val):
    """
    解析 wecom-cli 单元格值。
    
    wecom-cli 返回的 cell 值格式：
    - 数字: 95.5
    - 字符串: "95.5%"
    - 数组(文本): [{"type": "text", "text": "内容"}]
    - 数组(选项): [{"id": "xxx", "style": 7, "text": "待开始"}]
    - 对象: {"text": "内容"}
    """
    if cell_val is None:
        return None
    if isinstance(cell_val, (int, float)):
        return float(cell_val)
    if isinstance(cell_val, str):
        cell_val = cell_val.strip().rstrip('%').strip()
        if not cell_val:
            return None
        try:
            return float(cell_val)
        except ValueError:
            return cell_val
    if isinstance(cell_val, list) and len(cell_val) > 0:
        item = cell_val[0]
        if isinstance(item, dict):
            txt = item.get('text', '')
            if txt:
                txt = txt.strip().rstrip('%').strip()
                try:
                    return float(txt)
                except ValueError:
                    return txt
            return item.get('value', item.get('link', None))
        return str(item)
    if isinstance(cell_val, dict):
        txt = cell_val.get('text', cell_val.get('value', ''))
        if txt:
            txt = str(txt).strip().rstrip('%').strip()
            try:
                return float(txt)
            except ValueError:
                return txt
        return None
    return str(cell_val)


def run_cli(cli_path: str, tool: str, params_json: str) -> dict:
    """执行 wecom-cli doc 命令并返回解析后的 JSON 结果"""
    cmd = [cli_path, 'doc', tool, params_json]
    log.debug(f"  ▶ wecom-cli doc {tool} '{params_json[:80]}...'")
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        output = r.stdout.strip() or r.stderr.strip()
        if not output:
            return {'errcode': -1, 'errmsg': '无输出'}

        # 清理 ANSI 转义码
        import re
        output = re.sub(r'\x1b\[[0-9;]*[a-zA-Z]', '', output)

        # 尝试解析 JSON
        try:
            return json.loads(output)
        except json.JSONDecodeError:
            # 有时候输出前后可能有非 JSON 内容，尝试提取
            start = output.find('{')
            end = output.rfind('}')
            if start >= 0 and end > start:
                try:
                    return json.loads(output[start:end+1])
                except:
                    pass
            return {'errcode': -1, 'errmsg': f'JSON解析失败: {output[:200]}'}

    except subprocess.TimeoutExpired:
        return {'errcode': -1, 'errmsg': '命令超时(120s)'}
    except Exception as e:
        return {'errcode': -1, 'errmsg': str(e)}


# ============================================================
# 模式3: CDP 浏览器自动化（备选方案，占位）
# ============================================================
def fetch_cdp(conn, config: dict, queue_id: Optional[str] = None):
    """通过 CDP 从浏览器提取数据（备选方案）"""
    mode = config.get('global', {}).get('collector_mode', 'excel')
    if mode != 'cdp':
        log.info("当前采集模式非 cdp，跳过")
        return 0

    log.warning("⚠️ CDP 采集模式正在开发中，敬请期待")
    log.info("   建议：使用 excel 或 wcom-api 模式")
    return 0


# ============================================================
# 状态查询 & 导出
# ============================================================
def show_status(conn, config: dict):
    """显示各队列数据状态"""
    queues = config.get('queues', [])
    print("\n" + "=" * 76)
    print(f"  {'队列':<14} │ {'子表':<18} │ {'记录数':>6} │ {'日期范围':>16} │ 来源")
    print("=" * 76)

    for q in queues:
        qid = q['id']
        dates = get_queue_dates(conn, qid)
        count = len(dates)
        range_str = f"{dates[0]} ~ {dates[-1]}" if count > 0 else "(空)"
        source = "✅ 已接入" if count > 0 else "⏳ 待接入"
        icon = q.get('icon', '?')
        name = q.get('name', '')
        sheet = q.get('sheet_name', '')[:16]
        print(f"  {icon:>2} {name:<12} │ {sheet:<18} │ {count:>6} │ {range_str:>16} │ {source}")

    # wecom-cli 状态
    print("-" * 76)
    cli_ok = check_wecom_init()
    cli_status = "✅ 已认证" if cli_ok else "❌ 未初始化"
    cli_path = get_wecom_cli_path()
    cli_ver = "?"
    if cli_path:
        rv = subprocess.run([cli_path, '--version'], capture_output=True, text=True, timeout=5)
        cli_ver = rv.stdout.strip() or rv.stderr.strip() or "?"
    print(f"  wecom-cli: {cli_status} (v{cli_ver})")
    print("=" * 76)


def export_html_with_data(conn, output_path: Optional[str] = None):
    """将数据库中的数据注入到 HTML 看板模板中"""
    out = output_path or str(DASHBOARD_HTML)
    all_data = export_as_json(conn)

    with open(DASHBOARD_HTML, 'r', encoding='utf-8') as f:
        html_content = f.read()

    data_script = '<script>\n// Auto-generated data\nconst DB_DATA = ' + json.dumps(all_data, ensure_ascii=False) + ';\n</script>'
    html_content = html_content.replace('</body>', data_script + '\n</body>')

    with open(out, 'w', encoding='utf-8') as f:
        f.write(html_content)

    log.info(f"✅ 已导出带数据的看板: {out}")


# ============================================================
# CLI 入口
# ============================================================
def main():
    args = sys.argv[1:]
    if not args or args[0] in ('-h', '--help'):
        print(__doc__)
        return

    config = load_config()
    conn = init_db()

    action = args[0]
    queue_arg = None
    if '--queue' in args:
        idx = args.index('--queue')
        queue_arg = args[idx + 1] if idx + 1 < len(args) else None

    if action == 'setup':
        success = setup_wecom(config)
        sys.exit(0 if success else 1)

    elif action == 'fetch':
        mode = config.get('global', {}).get('collector_mode', 'excel')
        if mode == 'wcom-api':
            count = fetch_wecom_api(conn, config, queue_arg)
        elif mode == 'cdp':
            count = fetch_cdp(conn, config, queue_arg)
        else:
            count = import_excel(conn, config, queue_arg)
        log.info(f"\n🎉 完成！共处理 {count} 条数据")
        show_status(conn, config)

    elif action == 'import':
        count = import_excel(conn, config, queue_arg)
        log.info(f"\n🎉 导入完成！{count} 条数据")

    elif action == 'status':
        show_status(conn, config)

    elif action == 'export-html':
        export_html_with_data(conn)

    elif action == 'export-json':
        data = export_as_json(conn)
        out_file = DATA_DIR / f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(out_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        log.info(f"✅ JSON 导出: {out_file}")

    else:
        log.error(f"未知命令: {action}. 支持: setup, fetch, import, status, export-html, export-json")

    conn.close()


if __name__ == '__main__':
    main()
